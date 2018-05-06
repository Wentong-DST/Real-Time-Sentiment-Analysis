import nltk
from openpyxl import load_workbook
import  os
import numpy as np
import cPickle
from datetime import datetime
import pandas as pd
from twokenize import simple_tokenize
from word2vec_twitter_model.word2vecReader import Word2Vec

def label_transfer(label):
    if label == 0:
        return [1, 0]
    else:
        return [0, 1]

def divide_dataset(filename, ratio, sample):
    if not os.path.exists(filename):
        print 'file %s does not exist. Please correct the name and try again.' % filename

    # wb = load_workbook(filename)
    # ws = wb.active
    # row = ws.max_row
    # print '%d rows in file %s' % (row, filename)

    parser = lambda date: pd.datetime.strptime(date[:20]+date[24:], '%c')
    columns = ['polarity', 'id', 'date', 'query', 'user', 'text']
    df = pd.read_csv(filename, names=columns, parse_dates=[2], date_parser=parser) # tweet id as index
    categories = len(set(df['polarity']))
    if categories == 2:
        df['polarity'] = df['polarity'].apply(lambda x: x/4.0)
    elif categories == 3:
        df['polarity'] = df['polarity'].apply(lambda x: x/2.0)

    texts = df['text']
    labels = df['polarity']
    # get index for negative and positive samples
    neg_idx = labels.index[labels==0]
    if categories == 2:
        pos_idx = labels.index[labels==1]
    elif categories == 3:
        pos_idx = labels.index[labels==2]

    num_neg = len(neg_idx)
    num_pos = len(pos_idx)
    train_num_neg = int(num_neg * ratio)
    train_num_pos = int(num_pos * ratio)

    train_idx = neg_idx.tolist()[:train_num_neg] + pos_idx.tolist()[:train_num_pos]
    test_idx = list(set(neg_idx.tolist() + pos_idx.tolist()) - set(train_idx))
    train_idx = train_idx[:int(len(train_idx)*sample)]
    test_idx = test_idx[:int(len(test_idx)*sample)]
    # shuffle train samples
    tmp = np.arange(len(train_idx))
    np.random.shuffle(tmp)
    train_idx = np.array(train_idx)[tmp]
    train_texts = texts[train_idx]
    train_labels = labels[train_idx]
    train_labels = np.array(map(lambda l: label_transfer(l), train_labels.tolist()))

    # shuffle test samples
    tmp = np.arange(len(test_idx))
    np.random.shuffle(tmp)
    test_idx = np.array(test_idx)[tmp]
    test_texts = texts[test_idx]
    test_labels = labels[test_idx]
    test_labels = np.array(map(lambda l: label_transfer(l), test_labels.tolist()))

    return (train_texts, train_labels), (test_texts, test_labels)

"""
def remove_repeated_letter(word):
    result = word[:2]
    i = 2
    while i < len(word):
        a, b, c = word[i-2], word[i-1], word[i]
        if b == a and c == b:
            i += 1
            while i < len(word):
                if word[i] != c:
                    result += word[i]
                    break
                else:
                    i += 1
        else:
            result += c
        i += 1
    return result

def tweet_preporcessing(tweet):
    # remove urls
    url_idx = 0
    while url_idx >= 0:
        url_idx = tweet.find('http')
        whitespace = tweet.find(' ', start=url_idx)
        tweet = tweet.replace(tweet[url_idx: whitespace], 'url')

    # remove emoticons
    tweet = tweet.lower().replace(':)', 'smile').replace(':-)','smile').replace(': )','smile').replace(':D','smile').replace('=)','smile').replace(':(','sadface').replace(':-(','sadface').replace(': (','sadface')

    # remove repeated letters
    words = nltk.word_tokenize(tweet)
    words = list(map(lambda word: remove_repeated_letter(word), words))

def WordEmbedding(words, embedding_method='word2vec'):
    '''
    return word embeddings of words
    :param words: list of words
    :param embedding_method: the methdo for word embedding, fasttext, GloVe, Word2Vec
    :return:
    '''
    if embedding_method == 'fasttext':
        pass
    elif embedding_method == 'glove':
        pass
    elif embedding_method == 'word2vec':
        model = Word2Vec.load('w2v.model')

    words = list(map(lambda w: model.wv[w], words))     # list of (#words, #dim)
    return words
"""

def w2v(word, model):
    try:
        return model[word]
    except:
        return np.zeros((400,))

def tweet_preprocessing(tweets, model):
    '''

    :param tweets: list of tweets
    :return:
    '''
    # tokenization and replace URL, NUMBERs and MENTION with special tokens
    tweets = list(map(lambda t: simple_tokenize(t), tweets))  # list: (#tweets, list_of_words)
    # embedding
    embeddings = list(map(lambda tweet: np.array(list(map(lambda w: w2v(w, model), tweet))), tweets))   # list: (#tweets, np.array(#words, #dim))
    # padding
    max_words = max(list(map(lambda t: len(t), tweets)))
    paddings = np.array(list(map(lambda x: np.pad(x, ((0, max_words - x.shape[0]), (0, 0)), 'constant', constant_values=(0, 0)), embeddings)))

    return paddings


if __name__ == '__main__':
    # loading data and divide
    filename = 'Sentiment140/training.1600000.processed.noemoticon.csv'
    train_set, test_set = divide_dataset(filename)

    # loading word2vec model
    model_path = 'word2vec_twitter_model/word2vec_twitter_model.bin'
    model = Word2Vec.load_word2vec_format(model_path, binary=True)
    train_set[0] = tweet_preprocessing(train_set[0], model)
